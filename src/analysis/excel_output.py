# -*- coding: utf-8 -*-
"""
Created on Tue Nov 15 21:40:43 2022

@author: marti
"""
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from tqdm import tqdm
from src.analysis.parameters import sheets_parameters
from src.analysis.parameters import letters
from openpyxl.utils.dataframe import dataframe_to_rows
#%%
def write_in_the_excel(table, sheet, element, cell):
    letter = letters()
    # j=0
    # for column_title in table.keys():
    #     row = initial_line
    #     cell = column['letter'][element][j] + str(row)
    #     sheet[cell] = column_title
    #     j = j + 1
    #values start in the 'initial line' (row 5 of the excel sheet), then depends of the number of scenarios
    initial_line = int(cell[1:]) + 1  
    for i in tqdm(range(table.shape[0])): # going through all the rows
        row = initial_line + i #update cell reference
        for j in range(table.shape[1]): # going though all the columns                        
            cell = letter[j] + str(row) #update cell reference, from left to right                
            if table.iloc[i,j] == None or str(table.iloc[i,j]) == 'nan': # add --- if the value in the cell is None
                value = '---'
            else:
                value = table.iloc[i,j]  # add the value from the dataframe, auxiliar variable just to write the cell
            try: 
                sheet[cell] = value # update cell value
            except:
                print(value)
                print(cell)
                raise
    return cell
   
def create_excel(topology_name, output_path, tables) : 
    # read the template, write the results and add important parameters from the network topology
    # finally save into a new excel workbook, according to the network topology and scenario name 

    # read the template and retrieve the sheets
    output_template = 'src/analysis/output_templates/output_template.xlsm'
    print('\n Opening excel template')
    wb = load_workbook(filename = output_template, read_only = False, keep_vba = True) 
    print(' > Done')

    sheet, cell, elements_by_type = sheets_parameters() 

    print('\n Now writing into the excel file')
    for scenario in tables:
        print('\n Scenario: ' + scenario)
        for element in tables[scenario]:
            print(' > Element: ' + element)
            # print(tables[scenario][element])
            # write the values in excel table per sheet, and save the last cell
            try:
                cell[sheet[element]] = write_in_the_excel(tables[scenario][element], wb[sheet[element]], element, cell[sheet[element]])
            except:
                print('\nProblem writing in the excel')
                print('>>>>>> Scenario : '+ scenario)
                print('>>>>>> Element : '+ element)
                raise


    #% Update Table reference, and here the cell is the last cell added i.e. bottom-right corner of each table  
    print('\n\n Updating tables references in excel...')
    # the order of the element in the 'number' dict  is important, as it dictate which table in on top when
    # they are written in the same sheet, e.g. generation includes gen & sgen 
    sheets_names = elements_by_type
    for sheet_ in sheets_names:
            if not cell[sheet_] == cell['initial']:
                wb[sheet_].tables[sheet_].ref = cell['initial'] + ':' + cell[sheet_]
            
    #%% data sheet
    data = pd.DataFrame()
    for scenario in tables.keys():
        # print(scenario)
        for element in tables[scenario]:
            # print(element)
            df_element = tables[scenario][element]
            df_element['element'] = element
            data = pd.concat([data, df_element])
            
    data = data.replace(np.nan, '---')
    #data= data.replace(None, '--')

    rows = dataframe_to_rows(df = data)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
             wb['Data'].cell(row=r_idx, column=c_idx, value=value)
             
             
             
             

    #%% save with the topology and scenarios names
    print('\n Closing workbook ...')
    file_name =  'results_' + topology_name + '.xlsm'
    
    wb.save(output_path + '/' + file_name)
    print('\n Done' )
    print('      > You can check the results with the path: ' + output_path ) 
    print('      > and the results were saved in : ' + file_name)     

    return 

#%%













